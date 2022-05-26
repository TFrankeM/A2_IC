from openpyxl import Workbook, load_workbook

# Criar uma planilha (book)
arquivo_excel = Workbook()
# Criar uma planilha no arquivo
arquivo_excel.create_sheet('Planilha 1')
# Selecionar uma página para trabalhar
planilha = arquivo_excel['Planilha 1']
# Adicionando itens à planilha

planilha.append(['Moeda', 'Sigla', 'Quantidade', 'Cotação'])

planilha.append(['Dolar', 'USD', '200', 'R$5,8'])
planilha.append(['Rúbia', 'RUB', '17', '0,2'])
planilha.append(['Remimbi','CNY', '280', '0,71'])
planilha.append(['Iene', 'JPY', '150', '0,037'])

# Salvar a planilha
arquivo_excel.save('Carteira de Investimentos.xlsx')

