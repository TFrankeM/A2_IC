
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Fill, Color, NamedStyle, Border, Side, Alignment
from openpyxl.chart import BarChart, PieChart, DoughnutChart, Reference 
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image
from copy import deepcopy
import qrcode

        # <<<< EXCEL >>>>

    # INÍCIO DO EXCEL
def criar_arquivo():
    # Criar uma arquivo de trabalho no Excel
    arquivo_excel = Workbook()
    # Selecionar a planilha já existente 'Sheet' para trabalhar)
    planilha = arquivo_excel['Sheet']
    # Renomear essa planilha
    planilha.title = 'Carteira'
    return arquivo_excel, planilha

    # APARÊNCIA DAS CÉLULAS E LETRAS
# Aparência para o fundo do excel
def aparencia_arquivo(planilha):
    fundo = PatternFill(start_color='DBF3D9',
                   end_color='DBF3D9',
                   fill_type='solid')
    planilha["A1"].fill = fundo
    for linha in planilha["A1:CC100"]:
        for celula in linha:
            celula.fill = fundo


# Aparência para os cabeçalhos das tabelas
def aparencia_cabecalho(celula):
    celula.font = Font(bold=True, size=15)
    bd = Side(style='thick', color="000000")
    celula.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    celula.alignment = Alignment(horizontal='center', vertical="center")
    celula.fill = PatternFill(start_color='BCECF0', end_color='BCECF0', fill_type='solid')

# Criando aparência para as informações das tabelas
def aparencia_tabela(celula):
    celula.font = Font(size=12)
    celula.alignment = Alignment(horizontal='center', vertical="center")
    celula.fill = PatternFill(start_color='BCECF0', end_color='BCECF0', fill_type='solid')


# Largura das células
def largura(planilha):
    planilha.column_dimensions['B'].width = 15
    planilha.column_dimensions['C'].width = 22
    planilha.column_dimensions['D'].width = 22
    planilha.column_dimensions['E'].width = 22
    planilha.column_dimensions['P'].width = 4
    planilha.column_dimensions['Q'].width = 22

    # TABELAS DAS MOEDAS E DAS AÇÕES

    # Plotando os cabeçalhos
def celulas_fixas(planilha, carteira):
    # Cabeçalho das tabelas
    cabecalho = ['MOEDA','QUANTIDADE','COTAÇÃO (R$)','TOTAL (R$)']

    # Cabeçalho de MOEDAS
    for posicao, titulo in enumerate(cabecalho):
        celula = planilha.cell(row=2, column=posicao+2) 
        celula.value = titulo
        # Aplicando aparência no cabeçalho de moedas    
        aparencia_cabecalho(celula)

    # cabeçalho de AÇÕES
    for posicao, titulo in enumerate(cabecalho):
        celula = planilha.cell(row=len(carteira[0])+5, column=posicao+2)
        celula.value = titulo
        if posicao == 0:
            celula.value = 'AÇÕES'
    # Aplicando aparência no cabeçalho de ações 
        aparencia_cabecalho(celula)


    # Plotando a tabela de MOEDAS
def Plotar_moedas(planilha, carteira, carteira_cotacoes):
    # Coluna de nome
    for posicao, item in enumerate(carteira[0]):
        planilha[f'B{posicao+3}'] = item
        # Aplicando aparência
        celula = planilha.cell(row=posicao+3, column=2)
        aparencia_tabela(celula)

    # Coluna de quantidade
    for posicao, item in enumerate(carteira[1]): 
        planilha[f'C{posicao+3}'] = item
        # Aplicando aparência
        celula = planilha.cell(row=posicao+3, column=3)
        aparencia_tabela(celula)

    # Coluna de cotação
    for posicao, item in enumerate(carteira_cotacoes[0]): 
        planilha[f'D{posicao+3}'] = round(item, 2)
        # Aplicando aparência
        celula = planilha.cell(row=posicao+3, column=4)
        aparencia_tabela(celula)

    # Coluna de valor acumulado
    for posicao, item in enumerate(Total_moeda(carteira, carteira_cotacoes)): 
        planilha[f'E{posicao+3}'] = round(item, 2)
        # Aplicando aparência
        celula = planilha.cell(row=posicao+3, column=5)
        aparencia_tabela(celula)


    # Plotando a tabela de AÇÕES
def Plotar_acoes(planilha, carteira, carteira_cotacoes):
    # Coluna de nome
    for posicao, item in enumerate(carteira[2]): 
        planilha[f'B{posicao+len(carteira[0])+6}'] = item
        # Aplicando aparência
        celula = planilha.cell(row=posicao+len(carteira[0])+6, column=2)
        aparencia_tabela(celula)
    
    # Coluna de quantidade
    for posicao, item in enumerate(carteira[3]): 
        planilha[f'C{posicao+len(carteira[0])+6}'] = item
        # Aplicando aparência
        celula = planilha.cell(row=posicao+len(carteira[0])+6, column=3)
        aparencia_tabela(celula)

    # Coluna de cotação
    for posicao, item in enumerate(carteira_cotacoes[1]): 
        planilha[f'D{posicao+len(carteira[0])+6}'] = round(item, 2)
        # Aplicando aparência
        celula = planilha.cell(row=posicao+len(carteira[0])+6, column=4)
        aparencia_tabela(celula)

    # Coluna de valor acumulado
    for posicao, item in enumerate(Total_acao(carteira, carteira_cotacoes)):
        planilha[f'E{posicao+len(carteira[0])+6}'] = round(item, 2)
        # Aplicando aparência
        celula = planilha.cell(row=posicao+len(carteira[0])+6, column=5)
        aparencia_tabela(celula)


            # GRÁFICOS

def grafico1(planilha, carteira):
        # 1º Gráfico - Quantidade de ações X Valor Acumulado
    grafico_1 = PieChart()
    grafico_1.style = 26
    labels = Reference(planilha, min_col = 2, min_row = len(carteira[0])+6, max_row = len(carteira[2])+len(carteira[0])+5)                   
    data = Reference(planilha, min_col = 5, min_row = len(carteira[0])+5, max_row = len(carteira[2])+len(carteira[0])+5)
    grafico_1.add_data(data, titles_from_data = True) 
    grafico_1.set_categories(labels)
    grafico_1.title = "Capital total em cada ação"

    planilha.add_chart(grafico_1, f"B{len(carteira[2])+len(carteira[0])+8}")


def grafico2(planilha, carteira, carteira_cotacoes):
        # 2º Gráfico - Quantidade de ações X Valor Total
    # 'Gambiarra': plotar os dados novamente em uma coluna distante do dashboard, mas agora com as colunas desejadas juntas:
    for posicao, item in enumerate(carteira[1]):
        planilha['BY2'].value = 'QUANTIDADE'
        planilha[f'BY{posicao+3}'] = item
    for posicao, item in enumerate(Total_moeda(carteira, carteira_cotacoes)):
        planilha['BZ2'].value = 'TOTAL (R$)'
        planilha[f'BZ{posicao+3}'] = round(item, 2)
    # Adicionando o gráfico
    grafico_2 = BarChart()
    grafico_2.type = "col"
    grafico_2.style = 26
    labels = Reference(planilha, min_col = 2, min_row = 3, max_row = len(carteira[0])+2)                   
    data = Reference(planilha, min_col = 77, min_row = 2, max_row = len(carteira[0])+2, max_col = 78)
    grafico_2.add_data(data, titles_from_data = True) 
    grafico_2.set_categories(labels)
    grafico_2.title = "Comparação entre a quantidade e o valor total de cada moeda"
    grafico_2.x_axis.title = 'Moeda'

    planilha.add_chart(grafico_2, f"G2")


def grafico3(planilha, carteira, carteira_cotacoes):
        # 3º Gráfico - Quantidade de ações X Valor Acumulado
    # Gambiarra: plotar os dados novamente em uma coluna distante do dashboard, mas agora com as colunas desejadas juntas:
    # Nome ativo
    for posicao, item in enumerate(carteira[0]): 
        planilha[f'CB{posicao+3}'] = item
    for posicao, item in enumerate(carteira[2]): 
        planilha[f'CB{posicao+len(carteira[0])+3}'] = item
    planilha['CC2'].value = 'Total (R$) do ativo'
    # Quantidade e total da carteira
    for posicao, item in enumerate(Total_moeda(carteira, carteira_cotacoes)): 
        planilha[f'CC{posicao+3}'] = round(item, 2)
        planilha[f'CD{posicao+3}'] = round(Total_carteira(planilha, carteira, carteira_cotacoes)-item,2)
    for posicao, item in enumerate(Total_acao(carteira, carteira_cotacoes)):
        planilha[f'CC{posicao+len(carteira[0])+3}'] = round(item, 2)
        planilha[f'CD{posicao+len(carteira[0])+3}'] = round(Total_carteira(planilha, carteira, carteira_cotacoes)-item,2)
        planilha['CD2'].value = 'Total outros ativos'
    # Adicionando o gráfico
    # Gráfico intermediário com barras verticais que recebe as especificaçãos
    grafico_intermediario = BarChart()
    grafico_intermediario.type = "col"
    grafico_intermediario.style = 26
    grafico_intermediario.y_axis.title = 'Porcentagem'
    grafico_intermediario.x_axis.title = 'Ativos'
    labels = Reference(planilha, min_col=80, min_row=3, max_row=len(carteira[0])+len(carteira[2])+3)
    data = Reference(planilha, min_col=81, max_col=82, min_row=2, max_row=len(carteira[0])+len(carteira[2])+2)
    grafico_intermediario.add_data(data, titles_from_data=True)
    grafico_intermediario.set_categories(labels)
    # Gráfico definitivo com barras hotizontais a ser plotado
    grafico_3 = deepcopy(grafico_intermediario)
    grafico_3.type = "bar"
    grafico_3.style = 13
    grafico_3.grouping = "percentStacked"
    grafico_3.overlap = 100
    grafico_3.title = "Porcentagem investida em cada ativo em comparação ao valor total da carteira"

    planilha.add_chart(grafico_3, "G16")


       # VALOR ACUMULADO

        # Das MOEDAS
def Total_moeda(carteira, carteira_cotacoes):
    # Como o tamanho das listas 'carteira[1]' e 'carteira_cotacoes[0]' é o mesmo e elas são simétricas, podemos atribuir a
    # posição de uma lista na outra e encontrar o total de dinheiro, em reais investido em cada moeda
    valor_acum_moeda = []
    for posição, quantidade in enumerate(carteira[1]):
        valor_acum_moeda.append(quantidade * carteira_cotacoes[0][posição])
    return valor_acum_moeda


        # Das AÇÕES
def Total_acao(carteira, carteira_cotacoes):
    # Para ações é o mesmo caso de moedas
    valor_acum_acao = []
    for posição, quantidade in enumerate(carteira[3]):
        valor_acum_acao.append(quantidade * carteira_cotacoes[1][posição])
    return valor_acum_acao


        # VALOR TOTAL DA CARTEIRA
def Total_carteira(planilha, carteira, carteira_cotacoes):
        # Valor da carteira
    valor_carteira = sum(Total_moeda(carteira, carteira_cotacoes)) + sum(Total_acao(carteira, carteira_cotacoes))
        # QrCode com a mensagem do valor da carteira
    img_valor_carteira = qrcode.make(f'Esta carteira contem R${valor_carteira:.2f}', box_size = 3.9)
    img_valor_carteira.save('Valor da Carteira.png')
        # Adicionar uma legenda
    planilha['Q2'] = 'VALOR DA CARTEIRA:'
    planilha['Q2'].font = Font(size=12, bold=True)
    planilha['Q2'].alignment = Alignment(horizontal='center', vertical="center")

        # Adicionar a imagem no excel
    img = Image('Valor da Carteira.png')
    img.width = 160
    img.height = 160
    planilha.add_image(img, 'Q3')
    return(valor_carteira)

    # SALVAR O EXCEL
def salvar_excel(arquivo_excel, nome_excel):
    arquivo_excel.save(f'{nome_excel}.xlsx')


    # FUNÇÃO FINAL
# É ativada pela interface e, por sua vez, ativa as outras funções, criando o dashboard
def dashboard(carteira, carteira_cotacoes, nome_excel):
    arquivo_excel, planilha = criar_arquivo()

    aparencia_arquivo(planilha)
    celulas_fixas(planilha, carteira)
    largura(planilha)

    Total_moeda(carteira, carteira_cotacoes)
    Total_acao(carteira, carteira_cotacoes)
    Total_carteira(planilha, carteira, carteira_cotacoes)

    Plotar_moedas(planilha, carteira, carteira_cotacoes)
    Plotar_acoes(planilha, carteira, carteira_cotacoes)

    grafico1(planilha, carteira)
    grafico2(planilha, carteira, carteira_cotacoes)
    grafico3(planilha, carteira, carteira_cotacoes)

    salvar_excel(arquivo_excel, nome_excel)